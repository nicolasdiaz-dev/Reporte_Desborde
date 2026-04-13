import json
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
SPREADSHEET_ID = "1wtXzHb0Jl8OZK0_K5BZpI4GdNcsbi9Iom_h5tUE4C4w"
SHEET_NAME = "DATOS INB"

# ── Conexión ──────────────────────────────────────────────────────────────────
creds = Credentials.from_service_account_file("credenciales.json", scopes=SCOPES)
client = gspread.authorize(creds)
sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

data = sheet.get("A1:AC")
headers = data[0] + ["COL_AA", "TURNO", "SKILL"]
df = pd.DataFrame(data[1:], columns=headers)

# ── Limpieza ──────────────────────────────────────────────────────────────────
df["FECHA"] = df["FECHA"].astype(str).str.strip()
df["ANI_TELEFONO"] = df["ANI_TELEFONO"].astype(str).str.strip()
df["TURNO"] = df["TURNO"].fillna("").astype(str).str.strip()
df["SKILL"] = df["SKILL"].fillna("").astype(str).str.strip()

df = df[df["FECHA"].str.match(r"^\d{8}$")]
df["FECHA"] = pd.to_datetime(df["FECHA"], format="%Y%m%d").dt.strftime("%d/%m/%Y")

df_entrante = df[df["DIRECCION"].str.strip().str.upper() == "ENTRANTE"].copy()

# ── TABLA POR AGENTE ──────────────────────────────────────────────────────────
df_raw = df_entrante.copy()

def extraer_codigo(usuario):
    partes = str(usuario).split("-")
    return partes[0].strip() if partes[0].strip().isdigit() else "0"

df_raw["COD_AGENTE"] = df_raw["USUARIO"].apply(extraer_codigo)

def skill_primario(series):
    vals = series[~series.isin(["", "0", "nan"])]
    return vals.mode()[0] if not vals.empty else "0"

agente_skill  = df_raw.groupby("COD_AGENTE")["SKILL"].agg(skill_primario)
agente_turno  = df_raw.groupby("COD_AGENTE")["TURNO"].agg(skill_primario)
agente_nombre = df_raw.groupby("COD_AGENTE")["USUARIO"].agg(lambda s: s.mode()[0])

# ── Override manual de TURNO ──────────────────────────────────────────────────
TURNO_MANUAL = {
    "1348": "TM",
    "7417": "TM",
    "7423": "TM",
    "7443": "TM",
    "9140": "TM",
    "9281": "TM",
    "9297": "TM",
    "9321": "TM",
    "9372": "TM",
}
agente_turno = agente_turno.copy()
for cod, turno in TURNO_MANUAL.items():
    agente_turno[cod] = turno

pivot_agente = df_raw.groupby(["COD_AGENTE", "FECHA"])["ANI_TELEFONO"] \
    .nunique().reset_index().rename(columns={"ANI_TELEFONO": "COUNT_UNIQUE"})

pivot_agente_wide = pivot_agente.pivot_table(
    index="COD_AGENTE",
    columns="FECHA",
    values="COUNT_UNIQUE",
    fill_value=0,
).reset_index()
pivot_agente_wide.columns.name = None

pivot_agente_wide["SKILL"]   = pivot_agente_wide["COD_AGENTE"].map(agente_skill)
pivot_agente_wide["TURNO"]   = pivot_agente_wide["COD_AGENTE"].map(agente_turno)
pivot_agente_wide["USUARIO"] = pivot_agente_wide["COD_AGENTE"].map(agente_nombre)

fechas_cols = sorted([c for c in pivot_agente_wide.columns
                      if c not in ["COD_AGENTE", "SKILL", "TURNO", "USUARIO"]])

pivot_agente_wide["SUMA_TOTAL"] = pivot_agente_wide[fechas_cols].sum(axis=1)

aband = df_raw[df_raw["SUB_ESTADO"] == "RINGING"] \
    .groupby("COD_AGENTE")["ANI_TELEFONO"] \
    .nunique().reset_index().rename(columns={"ANI_TELEFONO": "ABAND"})

# ── Análisis abandonadas por hora y turno ────────────────────────────────────
df_aband_sin_agente = df_raw[
    (df_raw["SUB_ESTADO"] == "RINGING") & (df_raw["COD_AGENTE"] == "0")
].copy()
df_aband_sin_agente["HORA"] = pd.to_numeric(df_aband_sin_agente["HORA"], errors="coerce")
df_aband_sin_agente = df_aband_sin_agente.dropna(subset=["HORA"])
df_aband_sin_agente["TURNO_HORA"] = df_aband_sin_agente["HORA"].apply(
    lambda h: "TM" if h < 15 else "TT"
)

aband_hora_fecha = (
    df_aband_sin_agente
    .groupby(["TURNO_HORA", "HORA", "FECHA"])["ANI_TELEFONO"]
    .nunique().reset_index().rename(columns={"ANI_TELEFONO": "ABAND"})
)
aband_hora_pivot = aband_hora_fecha.pivot_table(
    index=["TURNO_HORA", "HORA"], columns="FECHA", values="ABAND", fill_value=0,
).reset_index()
aband_hora_pivot.columns.name = None

fechas_aband = sorted([c for c in aband_hora_pivot.columns if c not in ["TURNO_HORA", "HORA"]])
aband_hora_pivot["TOTAL"] = aband_hora_pivot[fechas_aband].sum(axis=1)
aband_hora_pivot = aband_hora_pivot.sort_values(["TURNO_HORA", "HORA"]).reset_index(drop=True)

subtotales = []
for turno in ["TM", "TT"]:
    mask = aband_hora_pivot["TURNO_HORA"] == turno
    fila = {"TURNO_HORA": turno, "HORA": "TOTAL"}
    for f in fechas_aband:
        fila[f] = aband_hora_pivot.loc[mask, f].sum()
    fila["TOTAL"] = aband_hora_pivot.loc[mask, "TOTAL"].sum()
    subtotales.append(fila)
aband_hora_pivot = pd.concat([aband_hora_pivot, pd.DataFrame(subtotales)], ignore_index=True)

pivot_agente_wide = pivot_agente_wide.merge(aband, on="COD_AGENTE", how="left")
pivot_agente_wide["ABAND"] = pivot_agente_wide["ABAND"].fillna(0).astype(int)

col_order = ["SKILL", "TURNO", "COD_AGENTE", "USUARIO", "ABAND"] + fechas_cols + ["SUMA_TOTAL"]
pivot_agente_wide = pivot_agente_wide[col_order]
pivot_agente_wide = pivot_agente_wide.sort_values(["SKILL", "TURNO", "COD_AGENTE"]).reset_index(drop=True)

# ── RESUMEN DESBORDE POR DÍA ──────────────────────────────────────────────────
total_dia      = pivot_agente_wide[fechas_cols].sum()
mask_inb       = pivot_agente_wide["SKILL"] == "INB"
total_inb      = pivot_agente_wide.loc[mask_inb, fechas_cols].sum()
total_desborde = pivot_agente_wide.loc[~mask_inb, fechas_cols].sum()

pct_inb      = (total_inb      / total_dia * 100).round(2)
pct_desborde = (total_desborde / total_dia * 100).round(2)

dias_con_actividad = total_dia[total_dia > 0]
promedio_desborde  = (total_desborde[dias_con_actividad.index] /
                      dias_con_actividad * 100).mean().round(2)

resumen_dia = pd.DataFrame({"Métrica": ["Total Día", "Total INB", "Desborde", "% INB", "% Desborde"]})
for fecha in fechas_cols:
    resumen_dia[fecha] = [
        int(total_dia[fecha]), int(total_inb[fecha]), int(total_desborde[fecha]),
        pct_inb[fecha], pct_desborde[fecha],
    ]
resumen_dia["Promedio"] = [
    "", "", "",
    (total_inb[dias_con_actividad.index] / dias_con_actividad * 100).mean().round(2),
    promedio_desborde,
]

# ── HUNTER ARGENTINA: cálculos diarios ───────────────────────────────────────
# QUEUED y RINGING se calculan desde df_entrante
queued_hunter  = (df_entrante[df_entrante["SUB_ESTADO"] == "QUEUED"]
                  .groupby("FECHA")["ANI_TELEFONO"].nunique())
ringing_hunter = (df_entrante[df_entrante["SUB_ESTADO"] == "RINGING"]
                  .groupby("FECHA")["ANI_TELEFONO"].nunique())

hunter_rows = []
for fecha in fechas_cols:
    total  = int(total_dia.get(fecha, 0))   # mismo origen que "Total Día" de la tabla INB
    desb   = int(total_desborde.get(fecha, 0))
    queued = int(queued_hunter.get(fecha, 0))
    ring   = int(ringing_hunter.get(fecha, 0))
    hunter_rows.append({
        "fecha":       fecha,
        "total":       total,
        "desborde":    desb,
        "pct_desb":    round(desb   / total * 100, 1) if total else 0.0,
        "pct_queued":  round(queued / total * 100, 1) if total else 0.0,
        "pct_ringing": round(ring   / total * 100, 1) if total else 0.0,
    })

hunter_data_json = json.dumps(hunter_rows)

# ── Export a Excel ────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font

output_file = "reporte_desborde_inb.xlsx"
AMARILLO_FILL   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
AMARILLO_FONT   = Font(bold=True, color="000000")
AZUL_TM_FILL    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
NARANJA_TT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUBTOTAL_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SUBTOTAL_FONT   = Font(bold=True)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    pivot_agente_wide.to_excel(writer, sheet_name="Por Agente", index=False)
    resumen_dia.to_excel(writer, sheet_name="Resumen Desborde por Día", index=False)
    aband_hora_pivot.to_excel(writer, sheet_name="Abandonadas por Hora", index=False)

    ws = writer.sheets["Por Agente"]
    n_cols = len(pivot_agente_wide.columns)
    for row_idx, cod in enumerate(pivot_agente_wide["COD_AGENTE"], start=2):
        if str(cod) == "0":
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = AMARILLO_FILL
                cell.font = AMARILLO_FONT
            break

    ws2 = writer.sheets["Abandonadas por Hora"]
    n_cols2 = len(aband_hora_pivot.columns)
    for row_idx, (turno, hora) in enumerate(
        zip(aband_hora_pivot["TURNO_HORA"], aband_hora_pivot["HORA"]), start=2
    ):
        es_subtotal = str(hora) == "TOTAL"
        fill = SUBTOTAL_FILL if es_subtotal else (AZUL_TM_FILL if turno == "TM" else NARANJA_TT_FILL)
        font = SUBTOTAL_FONT if es_subtotal else Font()
        for col_idx in range(1, n_cols2 + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = font

print(f"Archivo generado: {output_file}")

# ── Export a HTML ─────────────────────────────────────────────────────────────
promedio_inb  = float((total_inb[dias_con_actividad.index] / dias_con_actividad * 100).mean().round(2))
promedio_desb = float(promedio_desborde)

total_dia_js  = [int(total_dia[f]) for f in fechas_cols]
total_inb_js  = [int(total_inb[f]) for f in fechas_cols]
total_desb_js = [int(total_desborde[f]) for f in fechas_cols]
pct_inb_js    = [float(pct_inb[f]) for f in fechas_cols]
pct_desb_js   = [float(pct_desborde[f]) for f in fechas_cols]

mask_tm = pivot_agente_wide["TURNO"] == "TM"
mask_tt = pivot_agente_wide["TURNO"] == "TT"
desb_tm_dia = pivot_agente_wide.loc[mask_tm & ~mask_inb, fechas_cols].sum()
desb_tt_dia = pivot_agente_wide.loc[mask_tt & ~mask_inb, fechas_cols].sum()

pct_desb_tm_js = [
    round(float(desb_tm_dia[f] / total_dia[f] * 100), 2) if total_dia[f] > 0 else 0.0
    for f in fechas_cols
]
pct_desb_tt_js = [
    round(float(desb_tt_dia[f] / total_dia[f] * 100), 2) if total_dia[f] > 0 else 0.0
    for f in fechas_cols
]

dias_activos_mask = [v > 0 for v in total_dia_js]
n_activos    = max(sum(dias_activos_mask), 1)
prom_desb_tm = round(sum(v for v, a in zip(pct_desb_tm_js, dias_activos_mask) if a) / n_activos, 2)
prom_desb_tt = round(sum(v for v, a in zip(pct_desb_tt_js, dias_activos_mask) if a) / n_activos, 2)

def color_desborde(val):
    if val <= 10:
        return "#22c55e"
    t = min((val - 10) / 30.0, 1.0)
    r = int(0xf5 + (0xef - 0xf5) * t)
    g = int(0x9e + (0x44 - 0x9e) * t)
    b = int(0x0b + (0x44 - 0x0b) * t)
    return f"#{r:02x}{g:02x}{b:02x}"

def color_inb(val, promedio):
    if val >= promedio * 1.15: return "#22c55e"
    if val >= promedio * 0.85: return "#f59e0b"
    return "#ef4444"

filas_html = ""
metricas       = ["Total Día", "Total INB", "Desborde", "% INB", "% Desborde"]
valores_filas  = [total_dia_js, total_inb_js, total_desb_js, pct_inb_js, pct_desb_js]
promedios_fila = ["", "", "", promedio_inb, promedio_desb]

for metrica, vals, prom in zip(metricas, valores_filas, promedios_fila):
    filas_html += f'<tr><td class="metrica-col">{metrica}</td>'
    for v in vals:
        if metrica == "% Desborde":
            bg = color_desborde(v)
            filas_html += f'<td style="background:{bg};color:#fff;font-weight:600">{v}%</td>'
        elif metrica == "% INB":
            bg = color_inb(v, promedio_inb)
            filas_html += f'<td style="background:{bg};color:#fff;font-weight:600">{v}%</td>'
        else:
            filas_html += f'<td>{v}</td>'
    if metrica in ("% INB", "% Desborde"):
        filas_html += f'<td class="prom-col">{prom}%</td>'
    else:
        filas_html += '<td class="prom-col">—</td>'
    filas_html += "</tr>\n"

header_fechas = "".join(f"<th>{f}</th>" for f in fechas_cols)

html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Hunter Argentina · Reporte INB</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', sans-serif; background: #0f172a; color: #e2e8f0; padding: 24px 28px; }}

  /* ── Encabezado Hunter ── */
  .hunter-header {{
    display: flex; justify-content: space-between; align-items: flex-start;
    flex-wrap: wrap; gap: 12px; margin-bottom: 4px;
  }}
  h1 {{ font-size: 1.5rem; font-weight: 700; color: #f1f5f9; }}
  .subtitle {{ color: #64748b; font-size: 0.82rem; margin-bottom: 18px; }}

  /* ── Selector de días ── */
  .day-row {{ display: flex; align-items: center; gap: 10px; flex-wrap: wrap; margin-bottom: 16px; }}
  .day-label {{ font-size: 0.7rem; color: #64748b; text-transform: uppercase; letter-spacing: .05em; white-space: nowrap; }}
  .pills {{ display: flex; gap: 5px; flex-wrap: wrap; }}
  .day-pill {{
    background: #1e293b; border: 1px solid #334155; border-radius: 20px;
    padding: 3px 12px; font-size: 0.72rem; color: #94a3b8; cursor: pointer; transition: all .15s;
  }}
  .day-pill:hover {{ border-color: #3b82f6; color: #e2e8f0; }}
  .day-pill.active {{ background: #3b82f6; border-color: #3b82f6; color: #fff; font-weight: 600; }}

  /* ── Tabla Hunter ── */
  .hunter-wrap {{ overflow-x: auto; margin-bottom: 16px; }}
  .hunter-wrap table {{ border-collapse: collapse; width: 100%; min-width: 560px; }}
  .hunter-wrap th {{
    background: #1e293b; padding: 9px 16px; font-size: 0.72rem; text-transform: uppercase;
    letter-spacing: .05em; color: #64748b; border-bottom: 2px solid #334155; white-space: nowrap; text-align: center;
  }}
  .hunter-wrap th:first-child {{ text-align: left; }}
  .hunter-wrap td {{
    padding: 10px 16px; font-size: 0.9rem; border-bottom: 1px solid #1a2540;
    text-align: center; white-space: nowrap; transition: background .15s;
  }}
  .hunter-wrap td:first-child {{ text-align: left; font-weight: 600; color: #cbd5e1; }}
  .hunter-wrap tr.selected td {{ background: #1e3a5f !important; }}
  .hunter-wrap tr.selected td:first-child {{ color: #93c5fd; }}
  .hunter-wrap tr:hover td {{ background: #1e293b88; }}

  /* ── WhatsApp btn ── */
  .wa-btn {{
    display: inline-flex; align-items: center; gap: 7px;
    background: #25d366; color: #fff; font-weight: 600; font-size: 0.82rem;
    border: none; border-radius: 8px; padding: 8px 16px; cursor: pointer;
    text-decoration: none; transition: background .2s; white-space: nowrap;
  }}
  .wa-btn:hover {{ background: #1fbb58; }}
  .wa-btn svg {{ width: 16px; height: 16px; fill: #fff; }}

  /* ── Toggle análisis ── */
  .toggle-section {{
    display: flex; align-items: center; gap: 10px; cursor: pointer;
    background: #1e293b; border: 1px solid #334155; border-radius: 10px;
    padding: 10px 16px; margin-top: 8px; color: #64748b; font-size: 0.8rem;
    width: 100%; text-align: left; transition: background .2s;
  }}
  .toggle-section:hover {{ background: #263348; color: #e2e8f0; }}
  .toggle-section .arr {{ transition: transform .3s; font-size: 0.7rem; }}
  .toggle-section.open .arr {{ transform: rotate(180deg); }}

  /* ── Análisis completo (colapsable) ── */
  #analisis-wrap {{ margin-top: 20px; }}

  /* KPI cards */
  .kpis {{ display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap; }}
  .kpi {{ background: #1e293b; border-radius: 12px; padding: 16px 22px; flex: 1; min-width: 150px; border: 1px solid #334155; }}
  .kpi .label    {{ font-size: 0.72rem; color: #94a3b8; text-transform: uppercase; letter-spacing: .05em; margin-bottom: 1px; }}
  .kpi .sublabel {{ font-size: 0.65rem; color: #475569; margin-bottom: 6px; }}
  .kpi .value    {{ font-size: 1.9rem; font-weight: 700; margin-bottom: 5px; transition: color .25s; }}
  .kpi .prom-line {{ font-size: 0.7rem; color: #475569; }}
  .kpi .prom-line b {{ color: #22c55e; font-weight: 600; }}

  /* Trend */
  .trend-card {{ background: #1e293b; border-radius: 12px; padding: 14px 18px; border: 1px solid #334155; margin-bottom: 16px; }}
  .trend-card h2 {{ font-size: 0.74rem; color: #64748b; margin-bottom: 10px; text-transform: uppercase; letter-spacing: .05em; display: flex; justify-content: space-between; align-items: center; }}
  .trend-card h2 em {{ font-style: normal; font-size: 0.68rem; color: #22c55e; text-transform: none; letter-spacing: 0; }}

  /* Toggle detalle INB */
  .toggle-btn {{
    display: flex; align-items: center; gap: 10px; cursor: pointer;
    background: #1e293b; border: 1px solid #334155; border-radius: 10px;
    padding: 10px 18px; margin-bottom: 20px; color: #64748b; font-size: 0.82rem;
    width: 100%; text-align: left; transition: background .2s;
  }}
  .toggle-btn:hover {{ background: #263348; color: #e2e8f0; }}
  .toggle-btn .arrow {{ transition: transform .3s; font-size: 0.72rem; }}
  .toggle-btn.open .arrow {{ transform: rotate(180deg); }}

  /* Leyenda */
  .leyenda {{ display: flex; gap: 16px; margin-bottom: 16px; font-size: 0.75rem; align-items: center; }}
  .dot {{ width: 9px; height: 9px; border-radius: 50%; display: inline-block; margin-right: 4px; }}

  /* Tabla INB */
  .tabla-wrap {{ overflow-x: auto; margin-bottom: 36px; }}
  .tabla-wrap table {{ border-collapse: collapse; width: 100%; min-width: 600px; }}
  .tabla-wrap th {{ background: #1e293b; padding: 9px 13px; font-size: 0.75rem; text-transform: uppercase; letter-spacing: .04em; color: #94a3b8; border-bottom: 2px solid #334155; white-space: nowrap; }}
  .tabla-wrap td {{ padding: 8px 13px; font-size: 0.85rem; border-bottom: 1px solid #1e293b; text-align: center; white-space: nowrap; }}
  .tabla-wrap tr:hover td {{ background: #1e293b55; }}
  .metrica-col {{ text-align: left !important; font-weight: 600; color: #cbd5e1; background: #1e293b; position: sticky; left: 0; }}
  .prom-col {{ background: #1e293b; font-weight: 700; color: #f1f5f9; }}

  /* Gráficos */
  .charts {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  .chart-card {{ background: #1e293b; border-radius: 12px; padding: 20px; border: 1px solid #334155; }}
  .chart-card h2 {{ font-size: 0.78rem; color: #64748b; margin-bottom: 14px; text-transform: uppercase; letter-spacing: .05em; }}
  @media (max-width: 900px) {{ .charts {{ grid-template-columns: 1fr; }} }}

  /* INB day selector */
  #inb-pills {{ display: flex; gap: 5px; flex-wrap: wrap; margin-bottom: 20px; }}
</style>
</head>
<body>

<!-- ════════════════════════════════════════════════════════════════
     HUNTER ARGENTINA — vista principal
════════════════════════════════════════════════════════════════ -->
<div class="hunter-header">
  <div>
    <h1>Hunter Argentina</h1>
    <p class="subtitle">Llamadas entrantes · Resumen diario</p>
  </div>
  <button class="wa-btn" onclick="shareWhatsApp()">
    <svg viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
      <path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347m-5.421 7.403h-.004a9.87 9.87 0 01-5.031-1.378l-.361-.214-3.741.982.998-3.648-.235-.374a9.86 9.86 0 01-1.51-5.26c.001-5.45 4.436-9.884 9.888-9.884 2.64 0 5.122 1.03 6.988 2.898a9.825 9.825 0 012.893 6.994c-.003 5.45-4.437 9.884-9.885 9.884m8.413-18.297A11.815 11.815 0 0012.05 0C5.495 0 .16 5.335.157 11.892c0 2.096.547 4.142 1.588 5.945L.057 24l6.305-1.654a11.882 11.882 0 005.683 1.448h.005c6.554 0 11.89-5.335 11.893-11.893a11.821 11.821 0 00-3.48-8.413z"/>
    </svg>
    Compartir por WhatsApp
  </button>
</div>

<!-- Selector de día -->
<div class="day-row">
  <span class="day-label">Día:</span>
  <div class="pills" id="hunter-pills"></div>
</div>

<!-- Tabla Hunter -->
<div class="hunter-wrap">
  <table id="hunter-table">
    <thead>
      <tr>
        <th>Fecha</th>
        <th>Total Llamadas</th>
        <th>Total Desborde</th>
        <th>% Desborde</th>
        <th>% QUEUED</th>
        <th>% RINGING</th>
      </tr>
    </thead>
    <tbody id="hunter-tbody"></tbody>
  </table>
</div>

<button class="toggle-section" id="btnHunterExpand" onclick="toggleHunterRows()">
  <span class="arr">▼</span>
  <span>Ver todos los días</span>
</button>

<!-- ════════════════════════════════════════════════════════════════
     ANÁLISIS INB — colapsable
════════════════════════════════════════════════════════════════ -->
<button class="toggle-section" id="btnAnalisis" onclick="toggleAnalisis()">
  <span class="arr">▼</span>
  <span>Ver análisis completo de desborde INB</span>
</button>

<div id="analisis-wrap" style="display:none">

  <!-- Selector de día INB -->
  <div class="day-row">
    <span class="day-label">Día:</span>
    <div id="inb-pills"></div>
  </div>

  <!-- KPIs -->
  <div class="kpis">
    <div class="kpi">
      <div class="label">% Desborde · TM</div>
      <div class="sublabel">Turno Mañana</div>
      <div class="value" id="valTM">—</div>
      <div class="prom-line">Promedio: <b>{prom_desb_tm}%</b></div>
    </div>
    <div class="kpi">
      <div class="label">% Desborde · TT</div>
      <div class="sublabel">Turno Tarde</div>
      <div class="value" id="valTT">—</div>
      <div class="prom-line">Promedio: <b>{prom_desb_tt}%</b></div>
    </div>
    <div class="kpi">
      <div class="label">% Desborde Total</div>
      <div class="sublabel">TM + TT</div>
      <div class="value" id="valTotal">—</div>
      <div class="prom-line">Promedio: <b>{promedio_desb}%</b></div>
    </div>
    <div class="kpi">
      <div class="label">Días analizados</div>
      <div class="sublabel">&nbsp;</div>
      <div class="value" style="color:#f59e0b">{len(fechas_cols)}</div>
    </div>
  </div>

  <!-- Trend chart -->
  <div class="trend-card">
    <h2>Evolución % desborde por turno <em>— promedio del período</em></h2>
    <div style="position:relative;height:150px">
      <canvas id="chartTrend"></canvas>
    </div>
  </div>

  <!-- Toggle detalle INB -->
  <button class="toggle-btn" id="btnDetalle" onclick="toggleDetalle()">
    <span class="arrow">▼</span>
    <span>Ver detalle completo (tabla + gráficos)</span>
  </button>

  <div id="detalle" style="display:none">
    <div class="leyenda">
      <span><span class="dot" style="background:#22c55e"></span>≤ 10%</span>
      <span><span class="dot" style="background:#f59e0b"></span>10 – 25%</span>
      <span><span class="dot" style="background:#ef4444"></span>&gt; 25%</span>
    </div>
    <div class="tabla-wrap">
      <table>
        <thead>
          <tr>
            <th style="text-align:left">Métrica</th>
            {header_fechas}
            <th>Promedio</th>
          </tr>
        </thead>
        <tbody>{filas_html}</tbody>
      </table>
    </div>
    <div class="charts">
      <div class="chart-card">
        <h2>% Desborde por día</h2>
        <canvas id="chartDesborde"></canvas>
      </div>
      <div class="chart-card">
        <h2>Volumen por día (INB vs Desborde)</h2>
        <canvas id="chartVolumen"></canvas>
      </div>
    </div>
  </div>

</div><!-- /analisis-wrap -->

<script>
// ── Datos Hunter ──────────────────────────────────────────────────────────────
const hunterData = {hunter_data_json};
const fechas     = hunterData.map(r => r.fecha);

// ── Datos INB ─────────────────────────────────────────────────────────────────
const pctDesb    = {pct_desb_js};
const pctInb     = {pct_inb_js};
const volInb     = {total_inb_js};
const volDesb    = {total_desb_js};
const promDesb   = {promedio_desb};
const promInb    = {promedio_inb};
const pctDesbTM  = {pct_desb_tm_js};
const pctDesbTT  = {pct_desb_tt_js};
const promDesbTM = {prom_desb_tm};
const promDesbTT = {prom_desb_tt};

// ── Gradiente color ───────────────────────────────────────────────────────────
function lerpHex(a, b, t) {{
  const ah = parseInt(a.slice(1), 16), bh = parseInt(b.slice(1), 16);
  return '#' + [16, 8, 0].map(sh => {{
    const c = (ah >> sh) & 0xff, d = (bh >> sh) & 0xff;
    return Math.round(c + (d - c) * t).toString(16).padStart(2, '0');
  }}).join('');
}}
function colorDesb(pct) {{
  if (pct <= 10) return '#22c55e';
  const t = Math.min((pct - 10) / 30, 1);
  return lerpHex('#f59e0b', '#ef4444', t);
}}

// ── Render tabla Hunter ───────────────────────────────────────────────────────
const tbody = document.getElementById('hunter-tbody');
hunterData.forEach((r, i) => {{
  const tr = document.createElement('tr');
  tr.dataset.idx = i;
  tr.innerHTML = `
    <td>${{r.fecha}}</td>
    <td>${{r.total.toLocaleString()}}</td>
    <td>${{r.desborde.toLocaleString()}}</td>
    <td style="font-weight:600;color:${{colorDesb(r.pct_desb)}}">${{r.pct_desb}}%</td>
    <td style="font-weight:600;color:${{colorDesb(r.pct_queued)}}">${{r.pct_queued}}%</td>
    <td style="font-weight:600;color:${{colorDesb(r.pct_ringing)}}">${{r.pct_ringing}}%</td>
  `;
  tr.style.cursor = 'pointer';
  tr.onclick = () => selectHunterDay(i);
  tbody.appendChild(tr);
}});

// ── Pills Hunter ──────────────────────────────────────────────────────────────
let hunterIdx = null;
let hunterExpanded = false;
const hunterPills = document.getElementById('hunter-pills');
fechas.forEach((f, i) => {{
  const btn = document.createElement('button');
  btn.className = 'day-pill';
  btn.textContent = f;
  btn.onclick = () => selectHunterDay(i);
  hunterPills.appendChild(btn);
}});

function updateHunterVisibility() {{
  document.querySelectorAll('#hunter-tbody tr').forEach((tr, i) => {{
    tr.classList.toggle('selected', i === hunterIdx);
    tr.style.display = (hunterExpanded || i === hunterIdx) ? '' : 'none';
  }});
}}

function selectHunterDay(idx) {{
  hunterIdx = (hunterIdx === idx) ? null : idx;
  document.querySelectorAll('#hunter-pills .day-pill').forEach((p, i) =>
    p.classList.toggle('active', i === hunterIdx)
  );
  updateHunterVisibility();
}}

function toggleHunterRows() {{
  hunterExpanded = !hunterExpanded;
  const btn = document.getElementById('btnHunterExpand');
  btn.classList.toggle('open', hunterExpanded);
  btn.querySelector('span:last-child').textContent =
    hunterExpanded ? 'Ver menos' : 'Ver todos los días';
  updateHunterVisibility();
}}

// Preseleccionar anteúltima fecha
selectHunterDay(fechas.length - 2);

// ── WhatsApp share ────────────────────────────────────────────────────────────
function shareWhatsApp() {{
  const idx = hunterIdx !== null ? hunterIdx : fechas.length - 2;
  const r = hunterData[idx];
  const msg =
    `📊 *Hunter Argentina* · ${{r.fecha}}\n` +
    `━━━━━━━━━━━━━━━━\n` +
    `📞 Total Llamadas: ${{r.total.toLocaleString()}}\n` +
    `↩️  Total Desborde: ${{r.desborde.toLocaleString()}}\n` +
    `📉 % Desborde: ${{r.pct_desb}}%\n` +
    `⏳ % QUEUED: ${{r.pct_queued}}%\n` +
    `🔔 % RINGING: ${{r.pct_ringing}}%`;
  window.open('https://wa.me/?text=' + encodeURIComponent(msg), '_blank');
}}

// ── Toggle análisis INB ───────────────────────────────────────────────────────
let analisisVisible = false;
let inbChartsInit = false;
function toggleAnalisis() {{
  analisisVisible = !analisisVisible;
  document.getElementById('analisis-wrap').style.display = analisisVisible ? 'block' : 'none';
  const btn = document.getElementById('btnAnalisis');
  btn.classList.toggle('open', analisisVisible);
  btn.querySelector('span:last-child').textContent =
    analisisVisible ? 'Ocultar análisis INB' : 'Ver análisis completo de desborde INB';
  if (analisisVisible && !inbChartsInit) {{
    inbChartsInit = true;
    initInbPills();
    initTrendChart();
  }}
}}

// ── Pills INB ─────────────────────────────────────────────────────────────────
let inbIdx = null;
function initInbPills() {{
  const pillsEl = document.getElementById('inb-pills');
  fechas.forEach((f, i) => {{
    const btn = document.createElement('button');
    btn.className = 'day-pill';
    btn.textContent = f;
    btn.onclick = () => selectInbDay(i);
    pillsEl.appendChild(btn);
  }});
  selectInbDay(fechas.length - 2);
}}

function selectInbDay(idx) {{
  inbIdx = (inbIdx === idx) ? null : idx;
  document.querySelectorAll('#inb-pills .day-pill').forEach((p, i) =>
    p.classList.toggle('active', i === inbIdx)
  );
  updateINBKPIs();
}}

function updateINBKPIs() {{
  const tm    = inbIdx !== null ? pctDesbTM[inbIdx] : null;
  const tt    = inbIdx !== null ? pctDesbTT[inbIdx] : null;
  const total = inbIdx !== null ? pctDesb[inbIdx]   : null;
  const setVal = (id, val) => {{
    const el = document.getElementById(id);
    if (val === null) {{ el.textContent = '—'; el.style.color = '#475569'; }}
    else {{ el.textContent = val + '%'; el.style.color = colorDesb(val); }}
  }};
  setVal('valTM', tm); setVal('valTT', tt); setVal('valTotal', total);
}}

// ── Trend chart ───────────────────────────────────────────────────────────────
let trendChart = null;
function initTrendChart() {{
  trendChart = new Chart(document.getElementById('chartTrend'), {{
    type: 'line',
    data: {{
      labels: fechas,
      datasets: [
        {{
          label: 'TM', data: pctDesbTM, borderColor: '#3b82f6',
          backgroundColor: 'transparent', pointBackgroundColor: pctDesbTM.map(colorDesb),
          pointBorderColor: '#0f172a', pointRadius: 4, pointHoverRadius: 6,
          tension: 0.35, borderWidth: 2,
        }},
        {{
          label: 'TT', data: pctDesbTT, borderColor: '#f97316',
          backgroundColor: 'transparent', pointBackgroundColor: pctDesbTT.map(colorDesb),
          pointBorderColor: '#0f172a', pointRadius: 4, pointHoverRadius: 6,
          tension: 0.35, borderWidth: 2,
        }},
        {{
          label: 'Promedio', data: Array(fechas.length).fill(promDesb),
          borderColor: '#22c55e', borderDash: [5, 4], borderWidth: 1.5,
          pointRadius: 0, tension: 0, fill: false,
        }},
        {{
          label: 'Límite 10%', data: Array(fechas.length).fill(10),
          borderColor: '#ef4444', borderDash: [3, 3], borderWidth: 1,
          pointRadius: 0, tension: 0, fill: false,
        }}
      ]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{
        legend: {{ labels: {{ color: '#64748b', font: {{ size: 10 }}, boxWidth: 16, padding: 10 }} }},
        tooltip: {{ callbacks: {{ label: ctx => ctx.dataset.label + ': ' + ctx.parsed.y + '%' }} }}
      }},
      scales: {{
        x: {{ ticks: {{ color: '#475569', font: {{ size: 9 }} }}, grid: {{ color: '#172033' }} }},
        y: {{ ticks: {{ color: '#475569', font: {{ size: 9 }}, callback: v => v + '%' }},
              grid: {{ color: '#1e293b' }}, beginAtZero: true }}
      }}
    }}
  }});
}}

// ── Toggle detalle INB ────────────────────────────────────────────────────────
let detalleVisible = false, detalleChartsInit = false;
function toggleDetalle() {{
  detalleVisible = !detalleVisible;
  document.getElementById('detalle').style.display = detalleVisible ? 'block' : 'none';
  const btn = document.getElementById('btnDetalle');
  btn.classList.toggle('open', detalleVisible);
  btn.querySelector('span:last-child').textContent =
    detalleVisible ? 'Ocultar detalle' : 'Ver detalle completo (tabla + gráficos)';
  if (detalleVisible && !detalleChartsInit) {{ detalleChartsInit = true; initDetailCharts(); }}
}}

function initDetailCharts() {{
  new Chart(document.getElementById('chartDesborde'), {{
    type: 'bar',
    data: {{
      labels: fechas,
      datasets: [
        {{ label: '% Desborde', data: pctDesb, backgroundColor: pctDesb.map(colorDesb), borderRadius: 6 }},
        {{ label: 'Promedio', data: Array(fechas.length).fill(promDesb), type: 'line',
           borderColor: '#22c55e', borderDash: [6, 3], borderWidth: 2, pointRadius: 0, fill: false }},
        {{ label: 'Límite 10%', data: Array(fechas.length).fill(10), type: 'line',
           borderColor: '#ef4444', borderDash: [3, 3], borderWidth: 1.5, pointRadius: 0, fill: false }}
      ]
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }},
                  tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + '%' }} }} }},
      scales: {{
        x: {{ ticks: {{ color: '#94a3b8', font: {{ size: 10 }} }}, grid: {{ color: '#1e293b' }} }},
        y: {{ ticks: {{ color: '#94a3b8', callback: v => v + '%' }}, grid: {{ color: '#334155' }}, beginAtZero: true }}
      }}
    }}
  }});

  new Chart(document.getElementById('chartVolumen'), {{
    type: 'bar',
    data: {{
      labels: fechas,
      datasets: [
        {{ label: 'INB', data: volInb, backgroundColor: '#3b82f6', borderRadius: 4, stack: 'vol' }},
        {{ label: 'Desborde', data: volDesb, backgroundColor: '#ef444499', borderRadius: 4, stack: 'vol' }}
      ]
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ labels: {{ color: '#94a3b8', font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ ticks: {{ color: '#94a3b8', font: {{ size: 10 }} }}, grid: {{ color: '#1e293b' }}, stacked: true }},
        y: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: '#334155' }}, stacked: true, beginAtZero: true }}
      }}
    }}
  }});
}}
</script>
</body>
</html>
"""

html_file = "reporte_desborde_inb.html"
with open(html_file, "w", encoding="utf-8") as f:
    f.write(html)

print(f"HTML generado: {html_file}")
